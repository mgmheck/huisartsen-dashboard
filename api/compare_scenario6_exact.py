#!/usr/bin/env python3
"""
Vergelijk EXACTE scenario 6 waarden uit Excel kolom H (scen6_fte_midden_a) met ons model
"""

import sys
from pathlib import Path
import openpyxl

# Add current directory to path
sys.path.insert(0, str(Path(__file__).parent))

# Import calculator
from scenario_model import VraagCalculator

EXCEL_FILE = Path("/Users/mgmheck/Library/CloudStorage/OneDrive-Capaciteitsorgaan/040 - 049 HA/047 Capaciteitsplan/FINAL---Variant 3-9_3. Resultaat_raming_3.2_KHA_Versie 03092025_fte-gelijk-2035-14347---FINAL kopie.xlsx")

print("=" * 100)
print("🔬 EXACTE SCENARIO 6 VERGELIJKING")
print("=" * 100)

# Laad Excel workbook
print("\n📊 Laden Excel data...")
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
sheet = wb['Data vraag hoofdmodel']

# Vind kolom H (scen6_fte_midden_a)
print("🔍 Zoeken naar kolom 'scen6_fte_midden_a'...")

# Headers staan in rij 1
headers = {}
for col_idx in range(1, sheet.max_column + 1):
    cell = sheet.cell(1, col_idx)
    if cell.value:
        headers[str(cell.value).strip()] = col_idx

if 'scen6_fte_midden_a' not in headers:
    print("❌ Kolom 'scen6_fte_midden_a' niet gevonden!")
    sys.exit(1)

scen6_col = headers['scen6_fte_midden_a']
jaar_col = headers['jaar']

print(f"✅ Gevonden in kolom {openpyxl.utils.get_column_letter(scen6_col)}")

# Lees scenario 6 waarden voor 2025-2030
excel_scen6 = {}
for row_idx in range(2, sheet.max_row + 1):
    jaar = sheet.cell(row_idx, jaar_col).value
    if jaar and 2025 <= jaar <= 2030:
        scen6_value = sheet.cell(row_idx, scen6_col).value
        excel_scen6[jaar] = scen6_value

print(f"\n📋 Excel Scenario 6 waarden (kolom H - scen6_fte_midden_a):")
print("-" * 100)
for jaar in sorted(excel_scen6.keys()):
    print(f"   {jaar}: {excel_scen6[jaar]:>12,.2f} FTE")

# Bereken met ons model
print(f"\n🧮 Berekenen met ONS model...")
vraag_calc = VraagCalculator({})

model_scen6 = {}
for jaar in range(2025, 2031):
    result = vraag_calc.bereken_scenario6_additief(jaar, variant='midden')
    model_scen6[jaar] = result['fte_vraag']

print(f"\n📋 Ons Model Scenario 6 waarden:")
print("-" * 100)
for jaar in sorted(model_scen6.keys()):
    print(f"   {jaar}: {model_scen6[jaar]:>12,.2f} FTE")

# Vergelijk
print(f"\n{'=' * 100}")
print("📊 VERGELIJKING")
print("=" * 100)
print(f"\n{'Jaar':<6} {'Excel scen6':<15} {'Model scen6':<15} {'Verschil':<15} {'% Verschil':<12}")
print("-" * 100)

verschillen = []
for jaar in range(2025, 2031):
    excel_val = excel_scen6[jaar]
    model_val = model_scen6[jaar]
    verschil = excel_val - model_val
    pct = (verschil / excel_val) * 100 if excel_val != 0 else 0

    verschillen.append(verschil)

    print(f"{jaar:<6} {excel_val:>12,.2f}    {model_val:>12,.2f}    {verschil:>12,.2f}    {pct:>10,.1f}%")

# Analyse patroon
print(f"\n{'=' * 100}")
print("🔍 PATROON ANALYSE")
print("=" * 100)

print(f"\nAbsolute verschillen per jaar:")
for i, (jaar, verschil) in enumerate(zip(range(2025, 2031), verschillen)):
    if i > 0:
        delta = verschil - verschillen[i-1]
        print(f"   {jaar}: {abs(verschil):>10,.2f} FTE (Δ: {delta:>+10,.2f})")
    else:
        print(f"   {jaar}: {abs(verschil):>10,.2f} FTE")

# Check lineariteit
if len(verschillen) > 2:
    deltas = [verschillen[i] - verschillen[i-1] for i in range(1, len(verschillen))]
    avg_delta = sum(deltas) / len(deltas)
    max_deviation = max(abs(d - avg_delta) for d in deltas)

    print(f"\n📐 Lineariteit check:")
    print(f"   Gemiddelde jaarlijkse toename: {avg_delta:,.2f} FTE")
    print(f"   Maximale afwijking van lineair: {max_deviation:,.2f} FTE")

    if max_deviation < 1.0:
        print(f"   ✅ PERFECT LINEAIR patroon (afwijking < 1 FTE)")
        print(f"   📝 Formule: Verschil ≈ {avg_delta:,.2f} × (jaar - 2025)")
    else:
        print(f"   ⚠️  Niet perfect lineair (afwijking {max_deviation:,.2f} FTE)")

print(f"\n{'=' * 100}")
print("✅ ANALYSE VOLTOOID")
print("=" * 100)
