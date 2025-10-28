#!/usr/bin/env python3
"""
DEEP DIVE: Waar komt de 280 FTE/jaar vandaan?
"""

import openpyxl
from pathlib import Path

EXCEL_FILE = Path("/Users/mgmheck/Library/CloudStorage/OneDrive-Capaciteitsorgaan/040 - 049 HA/047 Capaciteitsplan/FINAL---Variant 3-9_3. Resultaat_raming_3.2_KHA_Versie 03092025_fte-gelijk-2035-14347---FINAL kopie.xlsx")

print("=" * 100)
print("🔬 DEEP DIVE: 280 FTE/JAAR MYSTERY")
print("=" * 100)

# STAP 1: Laad Excel MET formules (data_only=False)
print("\n📊 STAP 1: Laden Excel MET formules...")
wb_formulas = openpyxl.load_workbook(EXCEL_FILE, data_only=False)
sheet = wb_formulas['Data vraag hoofdmodel']

# Vind kolom H
headers = {}
for col_idx in range(1, min(30, sheet.max_column + 1)):
    cell = sheet.cell(1, col_idx)
    if cell.value:
        headers[str(cell.value).strip()] = col_idx

scen6_col = headers.get('scen6_fte_midden_a')
scen1_col = headers.get('scen1_fte_midden')
fte_totaal_col = headers.get('fte_totaal')

print(f"✅ Kolommen gevonden:")
print(f"   fte_totaal: {openpyxl.utils.get_column_letter(fte_totaal_col)}")
print(f"   scen1_fte_midden: {openpyxl.utils.get_column_letter(scen1_col)}")
print(f"   scen6_fte_midden_a: {openpyxl.utils.get_column_letter(scen6_col)}")

# Lees formules voor 2025-2027
print(f"\n🔍 STAP 2: Excel formules onderzoeken (2025-2027)")
print("-" * 100)

for row_idx in range(2, 5):  # 2025, 2026, 2027
    jaar_val = sheet.cell(row_idx, 2).value
    scen6_cell = sheet.cell(row_idx, scen6_col)
    scen1_cell = sheet.cell(row_idx, scen1_col)

    print(f"\n📅 Jaar {jaar_val} (Rij {row_idx}):")
    print(f"   Scenario 1 waarde: {scen1_cell.value}")
    print(f"   Scenario 1 formule: {scen1_cell.value if isinstance(scen1_cell.value, str) else 'Geen formule (directe waarde)'}")
    print(f"   Scenario 6 waarde: {scen6_cell.value}")

    # Check of de cel een formule heeft
    if hasattr(scen6_cell, 'value') and isinstance(scen6_cell.value, str) and scen6_cell.value.startswith('='):
        print(f"   Scenario 6 formule: {scen6_cell.value}")
    else:
        print(f"   Scenario 6 formule: Geen formule (directe waarde)")

# STAP 3: Laad Excel met ALLEEN waarden om te vergelijken
print(f"\n{'=' * 100}")
print(f"📊 STAP 3: Data analyse")
print("=" * 100)

wb_data = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
sheet_data = wb_data['Data vraag hoofdmodel']

print(f"\nVergelijking scenario 1 vs scenario 6 (Excel data):")
print("-" * 100)
print(f"{'Jaar':<6} {'FTE totaal':<15} {'Scen1':<15} {'Scen6':<15} {'Scen6-Scen1':<15}")
print("-" * 100)

for row_idx in range(2, 8):  # 2025-2030
    jaar = sheet_data.cell(row_idx, 2).value
    fte_tot = sheet_data.cell(row_idx, fte_totaal_col).value
    scen1 = sheet_data.cell(row_idx, scen1_col).value
    scen6 = sheet_data.cell(row_idx, scen6_col).value

    if scen1 and scen6:
        diff = scen6 - scen1
        print(f"{jaar:<6} {fte_tot:>12,.2f}    {scen1:>12,.2f}    {scen6:>12,.2f}    {diff:>12,.2f}")

# STAP 4: Analyseer "Data tabel invoer hoofdmodel" voor extra parameters
print(f"\n{'=' * 100}")
print(f"📋 STAP 4: Data tabel invoer hoofdmodel")
print("=" * 100)

if 'Data tabel invoer hoofdmodel' in wb_data.sheetnames:
    input_sheet = wb_data['Data tabel invoer hoofdmodel']

    print(f"\nSheet dimensies: {input_sheet.max_row} rijen × {input_sheet.max_column} kolommen")

    # Zoek naar scenario 6 gerelateerde parameters
    print(f"\nZoeken naar 'scenario 6' of 'scen6' gerelateerde parameters...")
    scen6_params = []

    for row_idx in range(1, min(120, input_sheet.max_row + 1)):
        for col_idx in range(1, min(15, input_sheet.max_column + 1)):
            cell = input_sheet.cell(row_idx, col_idx)
            if cell.value:
                value_str = str(cell.value).lower()
                if 'scen6' in value_str or ('scenario' in value_str and '6' in value_str):
                    scen6_params.append({
                        'row': row_idx,
                        'col': col_idx,
                        'value': cell.value
                    })

    if scen6_params:
        print(f"\n✅ Gevonden {len(scen6_params)} scenario 6 gerelateerde cellen:")
        for param in scen6_params[:20]:  # Toon eerste 20
            col_letter = openpyxl.utils.get_column_letter(param['col'])
            print(f"   {col_letter}{param['row']}: {param['value']}")

    # Zoek ook naar "additi" (additief)
    print(f"\nZoeken naar 'additief' gerelateerde parameters...")
    additief_params = []

    for row_idx in range(1, min(120, input_sheet.max_row + 1)):
        for col_idx in range(1, min(15, input_sheet.max_column + 1)):
            cell = input_sheet.cell(row_idx, col_idx)
            if cell.value:
                value_str = str(cell.value).lower()
                if 'additi' in value_str:
                    additief_params.append({
                        'row': row_idx,
                        'col': col_idx,
                        'value': cell.value
                    })

    if additief_params:
        print(f"\n✅ Gevonden {len(additief_params)} additief gerelateerde cellen:")
        for param in additief_params[:20]:
            col_letter = openpyxl.utils.get_column_letter(param['col'])
            print(f"   {col_letter}{param['row']}: {param['value']}")

    # Lees eerste 15 rijen en eerste 10 kolommen om structuur te zien
    print(f"\n📄 Eerste rijen van 'Data tabel invoer hoofdmodel':")
    print("-" * 100)

    for row_idx in range(1, min(16, input_sheet.max_row + 1)):
        row_data = []
        for col_idx in range(1, min(11, input_sheet.max_column + 1)):
            cell = input_sheet.cell(row_idx, col_idx)
            if cell.value:
                val_str = str(cell.value)[:25]
                row_data.append(f"{val_str:<25}")
            else:
                row_data.append(" " * 25)

        print(f"Rij {row_idx:2}: " + " | ".join(row_data[:5]))

else:
    print("\n❌ Sheet 'Data tabel invoer hoofdmodel' niet gevonden!")

# STAP 5: Zoek naar andere scenario 6 varianten (b, c, d, etc.)
print(f"\n{'=' * 100}")
print(f"🔍 STAP 5: Andere scenario 6 varianten?")
print("=" * 100)

sheet_vraag = wb_data['Data vraag hoofdmodel']
all_headers = []
for col_idx in range(1, sheet_vraag.max_column + 1):
    cell = sheet_vraag.cell(1, col_idx)
    if cell.value and 'scen6' in str(cell.value).lower():
        all_headers.append({
            'col': openpyxl.utils.get_column_letter(col_idx),
            'name': cell.value
        })

print(f"\n✅ Alle scenario 6 kolommen:")
for h in all_headers:
    print(f"   Kolom {h['col']}: {h['name']}")

print(f"\n{'=' * 100}")
print("✅ ANALYSE VOLTOOID")
print("=" * 100)
print(f"\n💡 VOLGENDE STAPPEN:")
print("   1. Controleer of Excel formules extra factoren bevatten")
print("   2. Bekijk 'Data tabel invoer hoofdmodel' voor scenario 6 parameters")
print("   3. Vergelijk niet-demografische parameters tussen varianten")
