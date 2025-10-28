#!/usr/bin/env python3
"""
Inspecteer Excel file om te bepalen welk scenario 'hoofdmodel' is
"""

import openpyxl
from pathlib import Path

EXCEL_FILE = Path("/Users/mgmheck/Library/CloudStorage/OneDrive-Capaciteitsorgaan/040 - 049 HA/047 Capaciteitsplan/FINAL---Variant 3-9_3. Resultaat_raming_3.2_KHA_Versie 03092025_fte-gelijk-2035-14347---FINAL kopie.xlsx")

print("=" * 80)
print("🔍 EXCEL SCENARIO ANALYSE")
print("=" * 80)

# Laad Excel workbook
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

print(f"\n📊 Bestand: {EXCEL_FILE.name}")
print(f"Aantal sheets: {len(wb.sheetnames)}\n")

# Toon alle sheet namen
print("📑 ALLE SHEETS:")
print("-" * 80)
for i, sheet_name in enumerate(wb.sheetnames, 1):
    sheet = wb[sheet_name]
    # Tel rijen met data
    max_row = sheet.max_row
    max_col = sheet.max_column
    print(f"{i:2}. {sheet_name:<50} ({max_row} rijen × {max_col} kolommen)")

# Analyseer "Data vraag hoofdmodel" sheet
print(f"\n{'=' * 80}")
print("🎯 ANALYSE: 'Data vraag hoofdmodel'")
print("=" * 80)

if 'Data vraag hoofdmodel' in wb.sheetnames:
    sheet = wb['Data vraag hoofdmodel']

    print(f"\n📋 Sheet dimensies: {sheet.max_row} rijen × {sheet.max_column} kolommen")

    # Lees eerste 20 rijen, eerste 10 kolommen om context te krijgen
    print(f"\n📄 Eerste rijen (eerste 10 kolommen):")
    print("-" * 80)

    for row_idx in range(1, min(21, sheet.max_row + 1)):
        row_data = []
        for col_idx in range(1, min(11, sheet.max_column + 1)):
            cell = sheet.cell(row_idx, col_idx)
            value = cell.value
            if value is not None:
                # Beperk lengte voor leesbaarheid
                value_str = str(value)[:30]
                row_data.append(f"{value_str:<30}")
            else:
                row_data.append(" " * 30)
        print(f"Rij {row_idx:2}: " + " | ".join(row_data[:5]))

    # Zoek naar scenario indicaties in de sheet
    print(f"\n🔎 Zoeken naar 'scenario' tekst in sheet...")
    scenario_mentions = []

    for row_idx in range(1, min(100, sheet.max_row + 1)):
        for col_idx in range(1, min(20, sheet.max_column + 1)):
            cell = sheet.cell(row_idx, col_idx)
            value = str(cell.value).lower() if cell.value else ""
            if 'scenario' in value or 'scen' in value:
                scenario_mentions.append({
                    'row': row_idx,
                    'col': col_idx,
                    'value': cell.value
                })

    if scenario_mentions:
        print(f"\n✅ Gevonden {len(scenario_mentions)} vermeldingen van 'scenario':")
        for mention in scenario_mentions[:10]:  # Toon eerste 10
            print(f"   Rij {mention['row']}, Kolom {mention['col']}: {mention['value']}")
    else:
        print(f"\n⚠️  Geen directe vermeldingen van 'scenario' gevonden in eerste 100 rijen")

    # Zoek naar kolom headers
    print(f"\n📊 Kolom headers (Rij 1):")
    print("-" * 80)
    for col_idx in range(1, min(21, sheet.max_column + 1)):
        cell = sheet.cell(1, col_idx)
        if cell.value:
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            print(f"   {col_letter}: {cell.value}")

else:
    print("\n❌ Sheet 'Data vraag hoofdmodel' niet gevonden!")

# Zoek ook naar andere vraag-gerelateerde sheets
print(f"\n{'=' * 80}")
print("🔍 VRAAG-GERELATEERDE SHEETS:")
print("=" * 80)

vraag_sheets = [name for name in wb.sheetnames if 'vraag' in name.lower() or 'scenario' in name.lower()]
if vraag_sheets:
    for sheet_name in vraag_sheets:
        print(f"\n📄 {sheet_name}")
        sheet = wb[sheet_name]

        # Toon eerste paar rijen
        print("   Eerste 3 rijen:")
        for row_idx in range(1, min(4, sheet.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(6, sheet.max_column + 1)):
                cell = sheet.cell(row_idx, col_idx)
                if cell.value:
                    row_data.append(str(cell.value)[:20])
            if row_data:
                print(f"      Rij {row_idx}: " + " | ".join(row_data))
else:
    print("Geen andere vraag-gerelateerde sheets gevonden")

print(f"\n{'=' * 80}")
print("✅ ANALYSE VOLTOOID")
print("=" * 80)
