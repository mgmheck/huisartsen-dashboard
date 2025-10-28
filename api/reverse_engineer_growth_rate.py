#!/usr/bin/env python3
"""
Reverse-engineer de groei-rate uit Excel data om te identificeren welke parameter
het verschil van ~280 FTE/jaar veroorzaakt tussen ons model en Excel scenario 6.
"""

import sys
from pathlib import Path
import openpyxl

# Add current directory to path
sys.path.insert(0, str(Path(__file__).parent))

# Import calculator
from scenario_model import VraagCalculator

EXCEL_FILE = Path("/Users/mgmheck/Library/CloudStorage/OneDrive-Capaciteitsorgaan/040 - 049 HA/047 Capaciteitsplan/FINAL---Variant 3-9_3. Resultaat_raming_3.2_KHA_Versie 03092025_fte-gelijk-2035-14347---FINAL kopie.xlsx")

print("=" * 100)
print("🔬 REVERSE ENGINEERING: GROEI-RATE ANALYSE")
print("=" * 100)

# Laad Excel workbook
print("\n📊 Laden Excel data...")
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
sheet = wb['Data vraag hoofdmodel']

# Vind kolommen
headers = {}
for col_idx in range(1, sheet.max_column + 1):
    cell = sheet.cell(1, col_idx)
    if cell.value:
        headers[str(cell.value).strip()] = col_idx

scen6_col = headers['scen6_fte_midden_a']
scen1_col = headers['scen1_fte_midden']
jaar_col = headers['jaar']

# Lees Excel scenario 6 waarden voor 2025-2030
excel_scen6 = {}
excel_scen1 = {}
for row_idx in range(2, sheet.max_row + 1):
    jaar = sheet.cell(row_idx, jaar_col).value
    if jaar and 2025 <= jaar <= 2030:
        excel_scen6[jaar] = sheet.cell(row_idx, scen6_col).value
        excel_scen1[jaar] = sheet.cell(row_idx, scen1_col).value

print(f"\n✅ Geladen Excel data voor jaren 2025-2030")

# Bereken met ons model
print(f"\n🧮 Berekenen met ONS model...")
vraag_calc = VraagCalculator({})

model_scen6 = {}
model_scen1 = {}
for jaar in range(2025, 2031):
    result6 = vraag_calc.bereken_scenario6_additief(jaar, variant='midden')
    result1 = vraag_calc.bereken_scenario1_demografisch(jaar, variant='midden')
    model_scen6[jaar] = result6['fte_vraag']
    model_scen1[jaar] = result1['fte_vraag']

# STAP 1: Bereken jaar-op-jaar groeipercentages uit Excel
print(f"\n{'=' * 100}")
print("📈 STAP 1: EXCEL GROEI-RATES (jaar-op-jaar)")
print("=" * 100)

excel_scen6_growth = {}
excel_scen1_growth = {}

print(f"\nScenario 6 groei (Excel):")
print("-" * 100)
print(f"{'Jaar':<6} {'FTE Start':>12} {'FTE Eind':>12} {'Absolute Δ':>12} {'Groei %':>12}")
print("-" * 100)

for jaar in range(2026, 2031):
    fte_start = excel_scen6[jaar-1]
    fte_eind = excel_scen6[jaar]
    absolute_delta = fte_eind - fte_start
    groei_pct = (absolute_delta / fte_start) * 100
    excel_scen6_growth[jaar] = groei_pct
    print(f"{jaar:<6} {fte_start:>12,.2f} {fte_eind:>12,.2f} {absolute_delta:>12,.2f} {groei_pct:>11,.3f}%")

print(f"\nScenario 1 groei (Excel):")
print("-" * 100)
print(f"{'Jaar':<6} {'FTE Start':>12} {'FTE Eind':>12} {'Absolute Δ':>12} {'Groei %':>12}")
print("-" * 100)

for jaar in range(2026, 2031):
    fte_start = excel_scen1[jaar-1]
    fte_eind = excel_scen1[jaar]
    absolute_delta = fte_eind - fte_start
    groei_pct = (absolute_delta / fte_start) * 100
    excel_scen1_growth[jaar] = groei_pct
    print(f"{jaar:<6} {fte_start:>12,.2f} {fte_eind:>12,.2f} {absolute_delta:>12,.2f} {groei_pct:>11,.3f}%")

# STAP 2: Bereken jaar-op-jaar groeipercentages uit ons model
print(f"\n{'=' * 100}")
print("📊 STAP 2: MODEL GROEI-RATES (jaar-op-jaar)")
print("=" * 100)

model_scen6_growth = {}
model_scen1_growth = {}

print(f"\nScenario 6 groei (Model):")
print("-" * 100)
print(f"{'Jaar':<6} {'FTE Start':>12} {'FTE Eind':>12} {'Absolute Δ':>12} {'Groei %':>12}")
print("-" * 100)

for jaar in range(2026, 2031):
    fte_start = model_scen6[jaar-1]
    fte_eind = model_scen6[jaar]
    absolute_delta = fte_eind - fte_start
    groei_pct = (absolute_delta / fte_start) * 100
    model_scen6_growth[jaar] = groei_pct
    print(f"{jaar:<6} {fte_start:>12,.2f} {fte_eind:>12,.2f} {absolute_delta:>12,.2f} {groei_pct:>11,.3f}%")

print(f"\nScenario 1 groei (Model):")
print("-" * 100)
print(f"{'Jaar':<6} {'FTE Start':>12} {'FTE Eind':>12} {'Groei %':>12}")
print("-" * 100)

for jaar in range(2026, 2031):
    fte_start = model_scen1[jaar-1]
    fte_eind = model_scen1[jaar]
    absolute_delta = fte_eind - fte_start
    groei_pct = (absolute_delta / fte_start) * 100
    model_scen1_growth[jaar] = groei_pct
    print(f"{jaar:<6} {fte_start:>12,.2f} {fte_eind:>12,.2f} {absolute_delta:>12,.2f} {groei_pct:>11,.3f}%")

# STAP 3: Vergelijk groei-rates tussen Excel en Model
print(f"\n{'=' * 100}")
print("🔍 STAP 3: GROEI-RATE VERSCHILLEN")
print("=" * 100)

print(f"\nScenario 6 groei-rate vergelijking:")
print("-" * 100)
print(f"{'Jaar':<6} {'Excel %':>12} {'Model %':>12} {'Verschil':>12} {'Verschil bps':>15}")
print("-" * 100)

groei_verschillen = []
for jaar in range(2026, 2031):
    excel_pct = excel_scen6_growth[jaar]
    model_pct = model_scen6_growth[jaar]
    verschil = excel_pct - model_pct
    verschil_bps = verschil * 100  # Basis points
    groei_verschillen.append(verschil)
    print(f"{jaar:<6} {excel_pct:>11,.3f}% {model_pct:>11,.3f}% {verschil:>11,.3f}% {verschil_bps:>14,.1f} bps")

gemiddeld_verschil = sum(groei_verschillen) / len(groei_verschillen)
print(f"\n📊 Gemiddeld groei-rate verschil: {gemiddeld_verschil:.3f}% ({gemiddeld_verschil*100:.1f} bps)")

# STAP 4: Bereken implied "missing factor" per jaar
print(f"\n{'=' * 100}")
print("🧬 STAP 4: IMPLIED MISSING FACTOR")
print("=" * 100)

print(f"\nAls we aannemen dat Excel's scenario 6 = Model's scenario 6 + extra factor...")
print("-" * 100)
print(f"{'Jaar':<6} {'Excel FTE':>12} {'Model FTE':>12} {'Missing FTE':>12} {'Impl. Factor':>15}")
print("-" * 100)

implied_factors = []
for jaar in range(2025, 2031):
    excel_fte = excel_scen6[jaar]
    model_fte = model_scen6[jaar]
    missing_fte = excel_fte - model_fte

    # Bereken implied factor: (Excel_FTE - fte_basis) / (Model_FTE - fte_basis)
    # Dit geeft de multiplicatieve factor op de groei
    fte_basis = 11447.30
    if jaar == 2025:
        implied_factor = 1.0  # Basis jaar
    else:
        excel_groei = excel_fte - fte_basis
        model_groei = model_fte - fte_basis
        if model_groei != 0:
            implied_factor = excel_groei / model_groei
        else:
            implied_factor = 0

    implied_factors.append(implied_factor)
    print(f"{jaar:<6} {excel_fte:>12,.2f} {model_fte:>12,.2f} {missing_fte:>12,.2f} {implied_factor:>14,.6f}")

# Check if implied factor is constant (would indicate multiplicative missing parameter)
if len(implied_factors) > 1:
    avg_factor = sum(implied_factors[1:]) / len(implied_factors[1:])  # Skip 2025
    max_deviation = max(abs(f - avg_factor) for f in implied_factors[1:])

    print(f"\n📐 Factor analyse:")
    print(f"   Gemiddelde implied factor: {avg_factor:.6f}")
    print(f"   Maximale afwijking: {max_deviation:.6f}")

    if max_deviation < 0.001:
        print(f"\n   ✅ IMPLIED FACTOR IS CONSTANT!")
        print(f"   📝 Excel scenario 6 = Model scenario 6 × {avg_factor:.6f}")
        print(f"   💡 Dit suggereert een MULTIPLICATIEVE missing parameter van ~{(avg_factor-1)*100:.2f}%")
    else:
        print(f"\n   ⚠️  IMPLIED FACTOR IS NIET CONSTANT")
        print(f"   📝 Dit suggereert COMPLEXERE missing factor (niet simpel multiplicatief)")

# STAP 5: Bereken wat de missing parameter per jaar zou zijn
print(f"\n{'=' * 100}")
print("🔎 STAP 5: DECOMPOSE MISSING GROWTH")
print("=" * 100)

print(f"\nOnze model componenten per jaar:")
print("-" * 100)

for jaar in range(2025, 2031):
    jaren_sinds_basis = jaar - 2025

    # Get scenario 1 result
    result1 = vraag_calc.bereken_scenario1_demografisch(jaar, variant='midden')
    scen1_fte = result1['fte_vraag']
    scen1_groei = (scen1_fte / fte_basis) - 1

    # Get scenario 6 result
    result6 = vraag_calc.bereken_scenario6_additief(jaar, variant='midden')
    scen6_fte = result6['fte_vraag']
    scen6_groei = (scen6_fte / fte_basis) - 1

    # Calculate niet_demo_factor from the relationship
    if scen1_groei != 0:
        niet_demo_factor = scen6_groei / scen1_groei
    else:
        niet_demo_factor = 1.0

    # Excel groei (implied)
    excel_fte = excel_scen6[jaar]
    fte_basis = 11447.30
    excel_groei = (excel_fte / fte_basis) - 1

    # Missing groei
    missing_groei = excel_groei - scen6_groei

    print(f"\nJaar {jaar} (t={jaren_sinds_basis}):")
    print(f"   Scenario 1 groei (demo+onv): {scen1_groei:>10,.4f} ({scen1_groei*100:>6,.2f}%)")
    print(f"   Niet-demo factor:            {niet_demo_factor:>10,.4f}")
    print(f"   Scenario 6 groei (model):    {scen6_groei:>10,.4f} ({scen6_groei*100:>6,.2f}%)")
    print(f"   Excel implied groei:         {excel_groei:>10,.4f} ({excel_groei*100:>6,.2f}%)")
    print(f"   Missing groei:               {missing_groei:>10,.4f} ({missing_groei*100:>6,.2f}%)")

    if jaren_sinds_basis > 0 and scen1_groei != 0:
        # Als Excel = Model + Missing, dan:
        # Excel_groei = niet_demo_factor * scen1_groei + X
        # X = Excel_groei - (niet_demo_factor * scen1_groei)
        # Als X = missing_parameter * jaren_sinds_basis * scen1_groei, dan:
        # missing_parameter = X / (jaren_sinds_basis * scen1_groei)

        implied_missing_param = missing_groei / (jaren_sinds_basis * scen1_groei)
        print(f"   Implied missing param/jaar:  {implied_missing_param:>10,.6f} ({implied_missing_param*100:>6,.3f}%)")

print(f"\n{'=' * 100}")
print("✅ ANALYSE VOLTOOID")
print("=" * 100)

print(f"\n💡 CONCLUSIE:")
print(f"   - Excel scenario 6 groeit consistent ~{gemiddeld_verschil:.3f}% sneller per jaar")
print(f"   - Dit resulteert in ~280 FTE verschil per jaar")
print(f"   - Check de 'implied missing param/jaar' waarden hierboven")
print(f"   - Vergelijk deze met niet-demografische parameters in CSV (epi, sociaal, etc.)")
